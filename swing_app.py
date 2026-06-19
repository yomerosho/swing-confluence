"""
swing_app.py — SwingConfluence Streamlit Dashboard
====================================================
Scan, view, and email 3-of-3 confluence swing setups.
"""

import streamlit as st
import pandas as pd
from datetime import datetime
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

from swing_scanner import (
    SwingScanner, ConfluenceSetup,
    ALL_TICKERS, INDICES_ETFS, MEGA_CAPS, SWING_NAMES,
)
from swing_html import build_swing_report, render_setup_card, PALETTE
from morning_briefing import render_morning_briefing


# ── Page config ───────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="SwingConfluence",
    page_icon="🎯",
    layout="wide",
)

# ── Secrets ───────────────────────────────────────────────────────────────────

def get_secret(section, key, default=""):
    try: return st.secrets[section][key]
    except: return default

ALPACA_KEY    = get_secret("alpaca", "key", "")
ALPACA_SECRET = get_secret("alpaca", "secret", "")
GMAIL_USER    = get_secret("gmail", "user", "")
GMAIL_PASS    = get_secret("gmail", "password", "")


# ── CSS ───────────────────────────────────────────────────────────────────────

st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;600;800&family=JetBrains+Mono:wght@300;400;600&display=swap');

/* Force dark theme globally */
html, body, .stApp, [class*="css"], .main, .block-container {{
  font-family: 'Syne', sans-serif !important;
  background: {PALETTE['bg']} !important;
  color: #ffffff !important;
}}

/* Main content area */
.main .block-container {{ background: {PALETTE['bg']} !important; padding-top: 2rem; }}
[data-testid="stAppViewContainer"] {{ background: {PALETTE['bg']} !important; }}
[data-testid="stHeader"] {{ background: {PALETTE['bg']} !important; }}
[data-testid="stToolbar"] {{ background: transparent !important; }}

/* TOP-RIGHT TOOLBAR BUTTONS (Share, edit, github, ⋮) — force light icons */
[data-testid="stToolbar"] button,
[data-testid="stToolbar"] a,
[data-testid="stToolbar"] svg {{
  color: #e0e8f0 !important;
  fill: #e0e8f0 !important;
  opacity: 1 !important;
}}
[data-testid="stToolbar"] button:hover,
[data-testid="stToolbar"] a:hover {{
  background: rgba(255,255,255,0.08) !important;
  color: #ffffff !important;
}}
header [data-testid="baseButton-headerNoPadding"] svg {{ color: #e0e8f0 !important; }}

/* Sidebar */
section[data-testid="stSidebar"] {{ background: #090c14 !important; border-right: 1px solid #1a2030; }}
section[data-testid="stSidebar"] * {{ color: #ffffff !important; }}
section[data-testid="stSidebar"] .stCheckbox label,
section[data-testid="stSidebar"] p,
section[data-testid="stSidebar"] span {{ color: #e0e8f0 !important; }}

/* Status messages */
.stStatus, [data-testid="stStatusWidget"] {{
  background: {PALETTE['card']} !important;
  color: #ffffff !important;
  border: 1px solid {PALETTE['border']} !important;
}}
.stStatus *, [data-testid="stStatusWidget"] * {{ color: #ffffff !important; }}
.stAlert {{ background: {PALETTE['card']} !important; color: #ffffff !important; }}
.stAlert * {{ color: #ffffff !important; }}

/* Progress bar messages */
.stMarkdown code, code {{
  background: {PALETTE['card_dark']} !important;
  color: #4af0c4 !important;
  padding: 4px 10px !important;
  border-radius: 6px !important;
  font-family: 'JetBrains Mono', monospace !important;
}}

/* Headings & text — high contrast */
h1, h2, h3, h4, h5, h6 {{ color: #ffffff !important; }}
p, span, label, div {{ color: #d4dce8; }}

/* Metric widgets */
[data-testid="metric-container"] {{
  background: {PALETTE['card']} !important;
  border: 1px solid {PALETTE['border']} !important;
  border-radius: 10px;
  padding: 14px 18px;
}}
[data-testid="metric-container"] label {{
  color: #a8b3c8 !important;
  font-family: 'JetBrains Mono', monospace !important;
  font-size: 0.7rem !important;
  text-transform: uppercase;
  letter-spacing: 0.1em;
}}
[data-testid="metric-container"] [data-testid="stMetricValue"] {{
  color: {PALETTE['brand']} !important;
  font-family: 'JetBrains Mono', monospace !important;
  font-weight: 700 !important;
}}

/* Text inputs — fix placeholder visibility */
.stTextInput input, .stTextArea textarea {{
  background: {PALETTE['card_dark']} !important;
  color: #ffffff !important;
  border: 1px solid {PALETTE['border']} !important;
  font-family: 'JetBrains Mono', monospace !important;
}}
.stTextInput input::placeholder,
.stTextArea textarea::placeholder {{
  color: #8090a0 !important;
  opacity: 1 !important;
}}
.stTextInput label, .stTextArea label {{ color: #d4dce8 !important; font-weight: 600; }}

/* Help tooltips (?) icon */
[data-testid="stTooltipIcon"] svg {{ color: #6a7a90 !important; fill: #6a7a90 !important; }}

/* Buttons */
.stButton > button {{
  background: linear-gradient(135deg, #3a1a6b, #1a0d4a) !important;
  color: {PALETTE['brand']} !important;
  border: none !important;
  border-radius: 6px;
  padding: 10px 20px;
  font-family: 'JetBrains Mono', monospace;
  font-weight: 600;
}}
.stButton > button:hover {{
  background: linear-gradient(135deg, #4a2a80, #2a0f60) !important;
  color: #ffffff !important;
}}

/* Captions */
.stCaption, small {{ color: #8a99b0 !important; }}

/* Horizontal divider */
hr {{ border-color: #1a2030 !important; }}

/* Checkbox color when checked */
.stCheckbox input:checked + div {{ background: {PALETTE['brand']} !important; }}

/* Custom-ticker chips — solid dark bg + bright text so symbols are readable.
   Scoped to the keyed rmtick_ wrappers so other buttons keep their look. */
[class*="st-key-rmtick_"] button {{
  background: {PALETTE['card']} !important;
  border: 1px solid {PALETTE['brand']} !important;
  border-radius: 999px !important;
  padding: 4px 12px !important;
  min-height: 0 !important;
  white-space: nowrap !important;
  font-family: 'JetBrains Mono', monospace !important;
  font-size: 0.8rem !important;
  font-weight: 700 !important;
}}
[class*="st-key-rmtick_"] button p,
[class*="st-key-rmtick_"] button div,
[class*="st-key-rmtick_"] button span {{
  color: {PALETTE['text']} !important;
}}
[class*="st-key-rmtick_"] button:hover {{
  background: #3a1f28 !important;
  border-color: {PALETTE['red']} !important;
}}
[class*="st-key-rmtick_"] button:hover p,
[class*="st-key-rmtick_"] button:hover div,
[class*="st-key-rmtick_"] button:hover span {{
  color: {PALETTE['red']} !important;
}}
</style>
""", unsafe_allow_html=True)

# ── Session state ─────────────────────────────────────────────────────────────

for k, v in {
    "setups":         [],
    "last_scan":      None,
    "scan_log":       [],
    "diagnostics":    [],
    "last_diagnostic": None,
    "custom_tickers":      [],   # persisted user-added tickers (chips)
    "custom_ticker_input": "",   # bound to the Add text box
    "_last_added":         [],   # for the "Added: X" confirmation
}.items():
    if k not in st.session_state:
        st.session_state[k] = v


# ── Custom-ticker callbacks (add / remove / clear) ──────────────────────────────

def _add_custom_tickers():
    """Parse the input box, append new valid tickers, then clear the box."""
    raw = st.session_state.get("custom_ticker_input", "")
    parsed = [
        t.strip().upper() for t in raw.replace(",", " ").split()
        if t.strip().isalpha() and 1 <= len(t.strip()) <= 6
    ]
    added = []
    for t in parsed:
        if t not in st.session_state.custom_tickers:
            st.session_state.custom_tickers.append(t)
            added.append(t)
    st.session_state._last_added = added
    st.session_state.custom_ticker_input = ""   # clear the box after adding

def _clear_custom_tickers():
    st.session_state.custom_tickers = []
    st.session_state._last_added = []
    st.session_state.custom_ticker_input = ""

def _remove_custom_ticker(ticker):
    if ticker in st.session_state.custom_tickers:
        st.session_state.custom_tickers.remove(ticker)
    st.session_state._last_added = [
        t for t in st.session_state.get("_last_added", []) if t != ticker
    ]

# ── Sidebar ───────────────────────────────────────────────────────────────────

with st.sidebar:
    st.markdown("## 🎯 SwingConfluence")
    st.markdown("*3-of-3 confluence · 4-of-4 ELITE with Strat · 1-3 day holds*")
    st.markdown("---")

    st.markdown("### 🎯 Scope")

    use_indices = st.checkbox("Indices/ETFs (3)", value=True)
    use_mega    = st.checkbox("Mega Caps (7)", value=True)
    use_swing   = st.checkbox("Swing Names (63)", value=True)

    selected = []
    if use_indices: selected += INDICES_ETFS
    if use_mega:    selected += MEGA_CAPS
    if use_swing:   selected += SWING_NAMES

    # ── Additional tickers (type → Add → removable chips) ────────────────
    st.text_input(
        "➕ Additional tickers",
        key="custom_ticker_input",
        placeholder="e.g. QCOM, SMCI, AVGO",
        help="Comma or space separated. Press Enter or click Add.",
        on_change=_add_custom_tickers,   # Enter also adds
    )

    ac1, ac2 = st.columns(2)
    ac1.button("➕ Add", key="add_custom_btn",
               on_click=_add_custom_tickers, use_container_width=True)
    ac2.button("🗑️ Clear all", key="clear_custom_btn",
               on_click=_clear_custom_tickers, use_container_width=True)

    if st.session_state.get("_last_added"):
        st.success(f"Added: {', '.join(st.session_state['_last_added'])}")

    # Current custom tickers as removable chips (click ✕ to drop one)
    custom_tickers = list(st.session_state.custom_tickers)
    if custom_tickers:
        st.caption("Current custom tickers:")
        chip_cols = st.columns(3)
        for i, t in enumerate(custom_tickers):
            chip_cols[i % 3].button(
                f"✕ {t}", key=f"rmtick_{t}",
                on_click=_remove_custom_ticker, args=(t,),
                use_container_width=True, help=f"Remove {t}",
            )

    # Merge into scan list (dedupe against the category selections)
    extra = [t for t in custom_tickers if t not in selected]
    selected += extra

    if extra:
        st.caption(f"Will scan {len(selected)} tickers (+{len(extra)} custom)")
    else:
        st.caption(f"Will scan {len(selected)} tickers")

    st.markdown("---")
    scan_btn  = st.button("🎯 Run Confluence Scan", type="primary")
    email_btn = st.button("📧 Email Results")

    if st.session_state.last_scan:
        st.caption(f"Last: {st.session_state.last_scan}")

    if ALPACA_KEY and ALPACA_SECRET:
        st.success("✅ Alpaca connected")
    else:
        st.error("❌ Alpaca keys missing")

    st.markdown(f"""
    <div style='margin-top:20px;font-size:0.7rem;color:#5a3a80;line-height:1.7;'>
    <b style='color:{PALETTE["brand"]};'>Conviction Tiers</b><br>
    <span style='color:#ff9f0a;'>⭐⭐⭐⭐⭐⭐⭐ ELITE — Strat confirmed</span><br>
    ⭐⭐⭐⭐⭐⭐ Daily + 4H aligned<br>
    ⭐⭐⭐⭐⭐ Daily signal<br>
    ⭐⭐⭐⭐ 4H signal only<br><br>
    <b style='color:{PALETTE["brand"]};'>Confluence Factors</b><br>
    Technical pattern (10 types)<br>
    GEX positioning supports<br>
    Whale flow ($500K+) agrees<br>
    <span style='color:#ff9f0a;'>⚡ The Strat combo / F2 / FTFC</span><br><br>
    <i style='color:#3a2a50;'>Educational only<br>Not financial advice</i>
    </div>""", unsafe_allow_html=True)


# ── Header ────────────────────────────────────────────────────────────────────

st.markdown(f"""
<div style='padding:8px 0 16px 0;'>
  <h1 style='font-family:Syne,sans-serif;font-size:2.2rem;font-weight:800;
             background:linear-gradient(90deg,{PALETTE["brand"]},#6a5aff,{PALETTE["green"]});
             -webkit-background-clip:text;-webkit-text-fill-color:transparent;
             margin:0;letter-spacing:-0.02em;'>🎯 SwingConfluence</h1>
  <p style='color:#5a3a80;font-family:JetBrains Mono,monospace;font-size:0.78rem;margin:4px 0 0 0;'>
    Daily + 4H pattern scan · GEX positioning · Whale flow · 1-3 day swings
  </p>
</div>
""", unsafe_allow_html=True)


# ── Scan ──────────────────────────────────────────────────────────────────────

from subscribers import get_emails as load_subscribers


def send_email(html, subject):
    if not GMAIL_USER or not GMAIL_PASS:
        return False, "Gmail credentials missing"
    subs = load_subscribers()
    if not subs:
        return False, "No subscribers"

    sent = 0
    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
            s.login(GMAIL_USER, GMAIL_PASS)
            for email in subs:
                msg = MIMEMultipart("alternative")
                msg["Subject"] = subject
                msg["From"]    = GMAIL_USER
                msg["To"]      = email
                msg.attach(MIMEText(html, "html"))
                s.send_message(msg)
                sent += 1
        return True, f"Sent to {sent} subscriber(s)"
    except Exception as e:
        return False, str(e)


def build_export_xlsx(setups):
    """Build a clean, readable xlsx for filtering and sorting setups."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── Palette ─────────────────────────────────────────────────
    BG_WHITE  = "FFFFFF"; BG_ALT    = "F7F8FC"; BG_HEADER = "1E2D4E"
    BG_ELITE  = "FFF3E0"; BG_MAX    = "EDE7F6"; BG_HIGH   = "E8F5E9"
    BG_CALL   = "E8F5E9"; BG_PUT    = "FFEBEE"
    FG_HEADER = "FFFFFF"; FG_ELITE  = "E65100"; FG_MAX    = "6A1B9A"
    FG_HIGH   = "2E7D32"; FG_CALL   = "1B5E20"; FG_PUT    = "B71C1C"
    FG_BODY   = "212121"; FG_MUTED  = "546E7A"; BORDER    = "CFD8DC"

    def _rr_color(v):
        if v >= 2.0: return "1B5E20"
        if v >= 1.5: return "2E7D32"
        if v >= 1.0: return "E65100"
        return "B71C1C"

    def _btm(color=BORDER):
        s = Side(style="thin", color=color)
        return Border(bottom=s)

    COLS = [
        ("Ticker",    10, "@"),   ("Dir",        6, "@"),
        ("★",          5, "0"),   ("Tier",        9, "@"),
        ("Spot",       9, "$#,##0.00"), ("Strike",  8, "$#,##0.00"),
        ("Expiry",    11, "@"),   ("Entry",       9, "$#,##0.00"),
        ("Stop",       9, "$#,##0.00"), ("T1",      9, "$#,##0.00"),
        ("T2",         9, "$#,##0.00"), ("R/R T1",  8, "0.00"),
        ("R/R T2",     8, "0.00"), ("OI",         13, "@"),
        ("Technical", 36, "@"),   ("GEX",         30, "@"),
        ("Whales",    34, "@"),   ("Strat",       30, "@"),
        ("FTFC",       6, "0"),   ("Daily",        7, "@"),
        ("4H",         5, "@"),   ("Whale $",     12, "$#,##0"),
        ("Support",    9, "$#,##0.00"), ("Resist",  9, "$#,##0.00"),
    ]

    tier_fg = {7: FG_ELITE, 6: FG_MAX, 5: FG_HIGH, 4: FG_MUTED}
    tier_bg = {7: BG_ELITE, 6: BG_MAX, 5: BG_HIGH, 4: BG_ALT}
    conv_lbl = {7: "ELITE", 6: "MAX", 5: "HIGH", 4: "MED"}

    wb = Workbook()
    ws = wb.active
    ws.title = "Setups"
    ws.sheet_view.showGridLines = False

    # ── Header row ───────────────────────────────────────────────
    ws.row_dimensions[1].height = 28
    hdr_fill  = PatternFill("solid", start_color=BG_HEADER)
    hdr_font  = Font(name="Calibri", bold=True, color=FG_HEADER, size=10)
    hdr_align = Alignment(horizontal="center", vertical="center")
    accent    = Side(style="medium", color="4FC3F7")

    for ci, (hdr, width, _) in enumerate(COLS, 1):
        c = ws.cell(row=1, column=ci, value=hdr)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = hdr_align
        c.border = Border(bottom=accent)
        ws.column_dimensions[get_column_letter(ci)].width = width

    # ── Data rows ────────────────────────────────────────────────
    for ri, s in enumerate(setups, 2):
        ws.row_dimensions[ri].height = 18
        row_bg   = BG_ALT if ri % 2 == 0 else BG_WHITE
        rr_t1    = getattr(s, "rr_t1",    s.risk_reward)
        rr_t2    = getattr(s, "rr_t2",    0.0)
        t1       = getattr(s, "target_t1", s.target)
        t2       = getattr(s, "target_t2", 0.0)
        ftfc_obj = getattr(s, "strat_ftfc", None)
        ftfc_sc  = ftfc_obj.score if ftfc_obj else 0
        tech     = " | ".join(f"{p.timeframe}: {p.pattern}" for p in s.patterns)
        strat    = getattr(s, "strat_summary", "")

        row_vals = [
            s.ticker, s.direction, s.conviction, conv_lbl.get(s.conviction, ""),
            round(s.spot, 2), round(s.strike, 2), s.expiry or "—",
            round(s.entry_above, 2), round(s.stop_below, 2),
            round(t1, 2), round(t2, 2) if t2 else "",
            rr_t1, rr_t2 if rr_t2 else "",
            getattr(s, "oi_quality", ""), tech,
            s.gex_summary, s.whale_summary, strat, ftfc_sc,
            "✓" if s.has_daily else "–",
            "✓" if s.has_4h   else "–",
            round(s.whale_premium, 0),
            s.support_level or "", s.resistance_level or "",
        ]

        for ci, (val, (_, _, fmt)) in enumerate(zip(row_vals, COLS), 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font   = Font(name="Calibri", size=9, color=FG_BODY)
            c.fill   = PatternFill("solid", start_color=row_bg)
            c.alignment = Alignment(vertical="center", horizontal="center")
            c.number_format = fmt
            c.border = _btm()

        # Ticker — left, bold
        tc = ws.cell(row=ri, column=1)
        tc.font = Font(name="Calibri", size=9, bold=True, color=FG_BODY)
        tc.alignment = Alignment(vertical="center", horizontal="left")

        # Direction
        dc = ws.cell(row=ri, column=2)
        dc.font = Font(name="Calibri", size=9, bold=True,
                       color=FG_CALL if s.direction == "CALL" else FG_PUT)
        dc.fill = PatternFill("solid", start_color=BG_CALL if s.direction == "CALL" else BG_PUT)

        # Stars + Tier
        for ci in (3, 4):
            cc = ws.cell(row=ri, column=ci)
            cc.font = Font(name="Calibri", size=9, bold=True,
                           color=tier_fg.get(s.conviction, FG_BODY))
            cc.fill = PatternFill("solid", start_color=tier_bg.get(s.conviction, row_bg))

        # R/R T1 bold colored
        rrc = ws.cell(row=ri, column=12)
        rrc.font = Font(name="Calibri", size=9, bold=True, color=_rr_color(rr_t1))

        # R/R T2 soft colored
        if rr_t2:
            rr2c = ws.cell(row=ri, column=13)
            rr2c.font = Font(name="Calibri", size=9, color=_rr_color(rr_t2))

        # Long text cols — left-align
        for ci in (15, 16, 17, 18):
            cc = ws.cell(row=ri, column=ci)
            cc.alignment = Alignment(vertical="center", horizontal="left")

        # Daily / 4H checkmarks
        for ci in (20, 21):
            cc = ws.cell(row=ri, column=ci)
            cc.font = Font(name="Calibri", size=9, bold=True,
                           color=FG_HIGH if cc.value == "✓" else FG_MUTED)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

    # ── Summary sheet ────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 12
    n = len(setups)

    summary = [
        ("SwingConfluence Export", None, None, True),
        (None, None, None, False),
        ("Metric", "Value", None, True),
        ("Total Setups",     f"=COUNTA(Setups!A2:A{n+1})", None, False),
        ("ELITE (7★)",       f"=COUNTIF(Setups!C2:C{n+1},7)", None, False),
        ("MAX (6★)",         f"=COUNTIF(Setups!C2:C{n+1},6)", None, False),
        ("HIGH (5★)",        f"=COUNTIF(Setups!C2:C{n+1},5)", None, False),
        ("Calls",            f'=COUNTIF(Setups!B2:B{n+1},"CALL")', None, False),
        ("Puts",             f'=COUNTIF(Setups!B2:B{n+1},"PUT")', None, False),
        (None, None, None, False),
        ("Best R/R T1",      f"=MAX(Setups!L2:L{n+1})", "0.00", False),
        ("Avg R/R T1",       f"=AVERAGE(Setups!L2:L{n+1})", "0.00", False),
        ("R/R ≥ 2.0",        f'=COUNTIF(Setups!L2:L{n+1},">=2")', None, False),
        ("R/R ≥ 1.5",        f'=COUNTIF(Setups!L2:L{n+1},">=1.5")', None, False),
    ]
    for ri, (label, val, fmt, bold) in enumerate(summary, 1):
        ws2.row_dimensions[ri].height = 20
        if label is None: continue
        la = ws2.cell(row=ri, column=1, value=label)
        if ri == 1:
            la.font = Font(name="Calibri", bold=True, size=12, color=BG_HEADER)
        elif ri == 3:
            la.font = Font(name="Calibri", bold=True, size=10, color=FG_HEADER)
            la.fill = PatternFill("solid", start_color=BG_HEADER)
        else:
            la.font = Font(name="Calibri", size=10, color=FG_BODY)
        if val is not None:
            vb = ws2.cell(row=ri, column=2, value=val)
            if ri == 3:
                vb.font = Font(name="Calibri", bold=True, size=10, color=FG_HEADER)
                vb.fill = PatternFill("solid", start_color=BG_HEADER)
            else:
                vb.font = Font(name="Calibri", size=10, bold=True, color=FG_HIGH)
            if fmt: vb.number_format = fmt

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

if scan_btn:
    if not ALPACA_KEY or not ALPACA_SECRET:
        st.error("⚠️ Configure Alpaca keys in Streamlit secrets first")
    else:
        scanner = SwingScanner(ALPACA_KEY, ALPACA_SECRET)
        with st.status("🎯 Running confluence scan...", expanded=True) as status:
            pb  = st.progress(0)
            stx = st.empty()

            def cb(pct, msg):
                pb.progress(pct)
                stx.markdown(f"`{msg}`")

            setups = scanner.scan_all(tickers=selected, progress_cb=cb)
            pb.empty(); stx.empty()

            st.session_state.setups    = setups
            st.session_state.last_scan = datetime.now().strftime("%H:%M:%S CT")

            if setups:
                status.update(label=f"✅ Found {len(setups)} confluence setup(s)!", state="complete")
            else:
                status.update(label="✅ Scan complete · No setups today", state="complete")


if email_btn:
    if not st.session_state.setups:
        st.warning("⚠️ Run a scan first")
    else:
        html = build_swing_report(st.session_state.setups, "Manual Scan")
        subj = f"🎯 SwingConfluence · {len(st.session_state.setups)} Setup(s) · {datetime.now().strftime('%b %d %H:%M CT')}"
        ok, msg = send_email(html, subj)
        if ok: st.success(f"✅ {msg}")
        else:  st.error(f"❌ {msg}")


# ── Tabs ──────────────────────────────────────────────────────────────────────

tab_scanner, tab_diagnostic, tab_morning = st.tabs([
    "🎯 Confluence Scanner",
    "🔬 Diagnostic Scan",
    "🌅 Morning Briefing",
])


# ════════════════════════════════════════════════
#  TAB 1 — CONFLUENCE SCANNER (main feature)
# ════════════════════════════════════════════════
with tab_scanner:
    if not st.session_state.setups and st.session_state.last_scan is None:
        st.markdown(f"""
        <div style='text-align:center;padding:80px 0;'>
          <div style='font-size:4rem;'>🎯</div>
          <h2 style='font-family:Syne,sans-serif;color:#5a3a80;margin-top:16px;'>3-of-3 Confluence · 4-of-4 ELITE Scanner</h2>
          <p style='font-family:JetBrains Mono,monospace;color:#3a2a50;font-size:0.9rem;'>
            Click <b style='color:{PALETTE["brand"]};'>Run Confluence Scan</b> in the sidebar
          </p>
          <p style='font-family:JetBrains Mono,monospace;color:#3a2a50;font-size:0.78rem;margin-top:24px;line-height:1.8;'>
            Scans Daily + 4H timeframes<br>
            Returns ONLY setups where ALL 3 factors align<br>
            Patterns · GEX positioning · Whale flow ≥ $500K
          </p>
          <p style='font-family:JetBrains Mono,monospace;color:#5a3a80;font-size:0.72rem;margin-top:30px;'>
            Not finding anything? Try the <b style='color:{PALETTE["brand"]};'>🔬 Diagnostic Scan</b> tab
          </p>
        </div>""", unsafe_allow_html=True)

    elif not st.session_state.setups:
        st.markdown(f"""
        <div style='background:{PALETTE["card"]};border:1px solid {PALETTE["border"]};
                    border-radius:12px;padding:40px;text-align:center;margin:20px 0;'>
          <div style='font-size:3rem;'>🔍</div>
          <h3 style='color:{PALETTE["text"]};font-family:monospace;margin-top:12px;'>No confluence setups today</h3>
          <p style='color:{PALETTE["text_dim"]};font-family:monospace;font-size:0.85rem;'>
            All tickers scanned. None meet the 3-of-3 threshold. (4-of-4 for ELITE)
          </p>
          <p style='color:{PALETTE["text_muted"]};font-family:monospace;font-size:0.78rem;margin-top:14px;'>
            Run the <b>🔬 Diagnostic Scan</b> tab to see which gate is filtering setups out.
          </p>
        </div>""", unsafe_allow_html=True)

    else:
        setups       = st.session_state.setups
        elite_count  = sum(1 for s in setups if s.conviction == 7)
        max_count    = sum(1 for s in setups if s.conviction == 6)
        high_count   = sum(1 for s in setups if s.conviction == 5)
        medium_count = sum(1 for s in setups if s.conviction == 4)
        call_count   = sum(1 for s in setups if s.direction == "CALL")
        put_count    = sum(1 for s in setups if s.direction == "PUT")

        c1, c2, c3, c4, c5, c6, c7 = st.columns(7)
        c1.metric("Total",    len(setups))
        c2.metric("ELITE 7★", elite_count)
        c3.metric("MAX 6★",   max_count)
        c4.metric("HIGH 5★",  high_count)
        c5.metric("MED 4★",   medium_count)
        c6.metric("Calls",    call_count)
        c7.metric("Puts",     put_count)

        # ── Export button ─────────────────────────────────────────
        from datetime import datetime as _dt
        export_xlsx = build_export_xlsx(setups)
        st.download_button(
            label="📥 Export to Excel",
            data=export_xlsx,
            file_name=f"SwingConfluence_{_dt.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="Download all setups as a filterable Excel spreadsheet",
        )

        st.markdown("---")

        for setup in setups:
            html = render_setup_card(setup)
            st.markdown(html, unsafe_allow_html=True)


# ════════════════════════════════════════════════
#  TAB 2 — DIAGNOSTIC SCAN
# ════════════════════════════════════════════════
with tab_diagnostic:
    st.markdown(f"""
    <div style='background:{PALETTE["card_dark"]};border:1px solid {PALETTE["border"]};
                border-left:3px solid {PALETTE["brand"]};border-radius:10px;
                padding:16px 20px;margin-bottom:20px;'>
      <div style='color:{PALETTE["brand"]};font-family:monospace;font-size:0.85rem;font-weight:700;margin-bottom:6px;'>
        🔬 Diagnostic Scanner
      </div>
      <div style='color:{PALETTE["text_dim"]};font-family:monospace;font-size:0.78rem;line-height:1.6;'>
        Shows where each ticker is being filtered out. Useful when the main scanner returns nothing.<br>
        For each ticker: <b>Patterns → GEX → Whales</b>. The "Blocked At" column tells you which gate fails.
      </div>
    </div>
    """, unsafe_allow_html=True)

    diag_btn = st.button("🔬 Run Diagnostic Scan", type="primary", key="diag_btn_main")

    health_btn = st.button("🩺 Quick Alpaca Health Check", key="health_btn")

    if health_btn:
        if not ALPACA_KEY or not ALPACA_SECRET:
            st.error("⚠️ Alpaca keys missing")
        else:
            import requests as _rq
            import json as _json

            st.markdown("### 🩺 Alpaca Data Quality Check")
            st.caption("Verifying which feeds work and inspecting the RAW options snapshot response.")

            headers = {
                "APCA-API-KEY-ID":     ALPACA_KEY,
                "APCA-API-SECRET-KEY": ALPACA_SECRET,
                "accept":              "application/json",
            }

            # 1. Stock quote test
            test_tickers = ["SPY", "AAPL"]
            results_rows = []

            with st.spinner("Probing stock quote endpoints..."):
                for ticker in test_tickers:
                    for feed in ["sip", "iex"]:
                        try:
                            r = _rq.get(
                                f"https://data.alpaca.markets/v2/stocks/{ticker}/quotes/latest",
                                headers=headers,
                                params={"feed": feed},
                                timeout=10,
                            )
                            status = r.status_code
                            if status == 200:
                                q = r.json().get("quote", {})
                                bid, ask = q.get("bp", 0), q.get("ap", 0)
                                mid_val = f"${(bid + ask) / 2:.2f}" if (bid and ask) else "—"
                                results_rows.append({
                                    "Ticker": ticker, "Feed": feed.upper(),
                                    "Status": status,
                                    "Bid": f"${bid:.2f}" if bid else "—",
                                    "Ask": f"${ask:.2f}" if ask else "—",
                                    "Mid": mid_val,
                                })
                            else:
                                results_rows.append({
                                    "Ticker": ticker, "Feed": feed.upper(),
                                    "Status": f"❌ {status}",
                                    "Bid": "—", "Ask": "—", "Mid": "—",
                                })
                        except Exception as e:
                            results_rows.append({
                                "Ticker": ticker, "Feed": feed.upper(),
                                "Status": f"❌ {type(e).__name__}",
                                "Bid": "—", "Ask": "—", "Mid": "—",
                            })

            st.markdown("#### Stock Quote Check")
            st.dataframe(pd.DataFrame(results_rows), use_container_width=True, hide_index=True)

            # 2. RAW OPTIONS SNAPSHOT INSPECTION — show fields actually populated
            st.markdown("#### Raw Options Snapshot — find which volume field has data")
            with st.spinner("Fetching SPY options snapshot..."):
                try:
                    from datetime import date as _date, timedelta as _td
                    today = _date.today()
                    end   = today + _td(days=7)

                    r = _rq.get(
                        "https://data.alpaca.markets/v1beta1/options/snapshots/SPY",
                        headers=headers,
                        params={
                            "limit": 5,
                            "feed": "indicative",
                            "expiration_date_gte": today.isoformat(),
                            "expiration_date_lte": end.isoformat(),
                        },
                        timeout=15,
                    )

                    if r.status_code == 200:
                        data = r.json()
                        snapshots = data.get("snapshots", {})
                        if snapshots:
                            # Show the first snapshot's full structure
                            first_symbol = list(snapshots.keys())[0]
                            first_snap   = snapshots[first_symbol]

                            st.markdown(f"**First contract: `{first_symbol}`**")
                            st.markdown("Inspect the fields below — we need to find which one actually has VOLUME:")

                            # Show all top-level keys
                            st.code(_json.dumps(list(first_snap.keys()), indent=2), language="json")

                            st.markdown("**Full snapshot JSON for this contract:**")
                            st.code(_json.dumps(first_snap, indent=2, default=str), language="json")

                            # Try to extract volume from all common locations
                            vol_check = {
                                "latestTrade.s":           first_snap.get("latestTrade", {}).get("s"),
                                "latestQuote.bs":          first_snap.get("latestQuote", {}).get("bs"),
                                "latestQuote.as":          first_snap.get("latestQuote", {}).get("as"),
                                "dailyBar.v":              first_snap.get("dailyBar", {}).get("v"),
                                "dailyBar.n":              first_snap.get("dailyBar", {}).get("n"),
                                "prevDailyBar.v":          first_snap.get("prevDailyBar", {}).get("v"),
                                "minuteBar.v":             first_snap.get("minuteBar", {}).get("v"),
                            }
                            st.markdown("**Volume candidates from this contract:**")
                            st.dataframe(
                                pd.DataFrame([
                                    {"Field": k, "Value": v if v is not None else "❌ MISSING"}
                                    for k, v in vol_check.items()
                                ]),
                                use_container_width=True, hide_index=True
                            )

                            # Aggregate across all 5 sample contracts
                            st.markdown("**Volume populated across 5 sample contracts:**")
                            agg = {k: 0 for k in vol_check.keys()}
                            for sym, snap in snapshots.items():
                                for k in agg:
                                    parts = k.split(".")
                                    val = snap.get(parts[0], {}).get(parts[1])
                                    if val is not None and val != 0:
                                        agg[k] += 1
                            st.dataframe(
                                pd.DataFrame([
                                    {"Field": k, "Contracts with non-zero data": f"{v}/{len(snapshots)}"}
                                    for k, v in agg.items()
                                ]),
                                use_container_width=True, hide_index=True
                            )

                        else:
                            st.error(f"No snapshots returned. Full response: {data}")
                    else:
                        st.error(f"Status {r.status_code}: {r.text[:500]}")

                except Exception as e:
                    st.error(f"Probe failed: {e}")

    if diag_btn:
        if not ALPACA_KEY or not ALPACA_SECRET:
            st.error("⚠️ Configure Alpaca keys in Streamlit secrets first")
        else:
            scanner = SwingScanner(ALPACA_KEY, ALPACA_SECRET)
            with st.status("🔬 Running diagnostic scan...", expanded=True) as status:
                pb  = st.progress(0)
                stx = st.empty()

                def diag_cb(pct, msg):
                    pb.progress(pct)
                    stx.markdown(f"`{msg}`")

                diags = scanner.diagnose_all(tickers=selected, progress_cb=diag_cb)
                pb.empty(); stx.empty()

                st.session_state.diagnostics    = diags
                st.session_state.last_diagnostic = datetime.now().strftime("%H:%M:%S CT")
                status.update(label=f"✅ Diagnosed {len(diags)} tickers", state="complete")

    # Display diagnostic results
    if st.session_state.diagnostics:
        diags = st.session_state.diagnostics

        # ── Aggregate stats ─────────────────────────────────────────────
        no_data       = sum(1 for d in diags if d.get("error") or not d.get("spot"))
        no_patterns   = sum(1 for d in diags if not d.get("daily_patterns") and not d.get("h4_patterns") and not d.get("error"))
        had_patterns  = len(diags) - no_data - no_patterns

        # Count where setups blocked
        blocked_gex          = 0
        blocked_whales       = 0
        passed_all           = 0
        passed_with_override = 0
        no_chain             = 0
        for d in diags:
            for r in d.get("results", []):
                if r["blocked_at"] == "gex":      blocked_gex    += 1
                elif r["blocked_at"] == "whales": blocked_whales += 1
                elif r["blocked_at"] == "passed": passed_all     += 1
                elif r["blocked_at"] == "passed_with_override":
                    passed_all           += 1
                    passed_with_override += 1
                elif r["blocked_at"] == "no_chain": no_chain     += 1

        st.markdown("### 📊 Gate Analysis Summary")

        c1, c2, c3, c4, c5, c6 = st.columns(6)
        c1.metric("Tickers", len(diags))
        c2.metric("Had Patterns", had_patterns)
        c3.metric("No Patterns", no_patterns)
        c4.metric("Blocked: GEX", blocked_gex)
        c5.metric("Blocked: Whales", blocked_whales)
        c6.metric("✅ Passed", passed_all, delta=f"{passed_with_override} via whale override" if passed_with_override else None)

        # Insight
        total_attempts = blocked_gex + blocked_whales + passed_all + no_chain
        if total_attempts > 0:
            primary_blocker = max(
                [("GEX", blocked_gex), ("Whales", blocked_whales),
                 ("Chain unavailable", no_chain), ("All passed", passed_all)],
                key=lambda x: x[1]
            )
            pct_blocked = primary_blocker[1] / total_attempts * 100

            color = PALETTE["red"] if primary_blocker[0] != "All passed" else PALETTE["green"]
            st.markdown(f"""
            <div style='background:{PALETTE["card"]};border-left:3px solid {color};
                        border-radius:8px;padding:14px 18px;margin:16px 0;'>
              <div style='color:{color};font-family:monospace;font-size:0.95rem;font-weight:700;'>
                🔍 Primary blocker: {primary_blocker[0]} ({primary_blocker[1]}/{total_attempts} = {pct_blocked:.0f}%)
              </div>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("---")

        # ── Per-ticker breakdown ────────────────────────────────────────
        st.markdown("### 🎯 Per-Ticker Breakdown")

        filter_opt = st.radio(
            "Show:",
            ["All tickers", "Only blocked at whales", "Only blocked at GEX", "Only with patterns", "Only passed all"],
            horizontal=True,
        )

        rows = []
        for d in diags:
            ticker = d["ticker"]
            spot   = d.get("spot", 0) or 0

            if d.get("error"):
                rows.append({
                    "Ticker": ticker, "Spot": "—",
                    "Daily Patterns": "—", "4H Patterns": "—",
                    "Direction": "—", "GEX": "—", "Whales": "—",
                    "Blocked At": f"❌ {d['error']}",
                    "Whale $": "—",
                })
                continue

            daily_count = len(d.get("daily_patterns", []))
            h4_count    = len(d.get("h4_patterns",    []))

            if not d.get("results"):
                # No patterns at all
                rows.append({
                    "Ticker": ticker, "Spot": f"${spot:.2f}",
                    "Daily Patterns": daily_count, "4H Patterns": h4_count,
                    "Direction": "—", "GEX": "—", "Whales": "—",
                    "Blocked At": "🚫 No patterns",
                    "Whale $": "—",
                })
                continue

            for r in d["results"]:
                rows.append({
                    "Ticker": ticker, "Spot": f"${spot:.2f}",
                    "Daily Patterns": daily_count, "4H Patterns": h4_count,
                    "Direction": r["direction"],
                    "GEX":    "✅" if r["gex_supports"] else "❌",
                    "Whales": "✅" if r["whale_supports"] else "❌",
                    "Blocked At": {
                        "patterns":              "🚫 No patterns",
                        "gex":                   "❌ GEX",
                        "whales":                "❌ Whales",
                        "no_chain":              "⚠️ No chain",
                        "passed":                "✅ PASSED",
                        "passed_with_override":  "✅ PASSED (whale override)",
                    }.get(r["blocked_at"], r["blocked_at"]),
                    "Whale $": f"${r['whale_top_premium']:,.0f}" if r['whale_top_premium'] else "—",
                })

        df_diag = pd.DataFrame(rows)

        # Apply filter
        if filter_opt == "Only blocked at whales":
            df_diag = df_diag[df_diag["Blocked At"].str.contains("Whales", na=False)]
        elif filter_opt == "Only blocked at GEX":
            df_diag = df_diag[df_diag["Blocked At"].str.contains("GEX", na=False)]
        elif filter_opt == "Only with patterns":
            df_diag = df_diag[(df_diag["Daily Patterns"] != "—") &
                              ((df_diag["Daily Patterns"] > 0) | (df_diag["4H Patterns"] > 0))]
        elif filter_opt == "Only passed all":
            df_diag = df_diag[df_diag["Blocked At"].str.contains("PASSED", na=False)]

        st.dataframe(df_diag, use_container_width=True, hide_index=True)

        # ── Detailed pattern + summary breakdown ─────────────────────────
        st.markdown("---")
        st.markdown("### 📋 Detailed Pattern Detections")
        st.caption("Expand each ticker to see exactly what patterns fired and why GEX/whales blocked it.")

        tickers_with_patterns = [d for d in diags
                                  if d.get("daily_patterns") or d.get("h4_patterns")]

        if not tickers_with_patterns:
            st.info("No patterns detected on any ticker. This means Gate 1 (pattern detection) is the bottleneck.")
        else:
            for d in tickers_with_patterns:
                ticker = d["ticker"]
                spot   = d.get("spot", 0)

                summary_emoji = "✅" if any(r["blocked_at"] == "passed" for r in d.get("results", [])) else "❌"

                with st.expander(f"{summary_emoji} {ticker} · ${spot:.2f} · {len(d.get('daily_patterns', []))}d + {len(d.get('h4_patterns', []))}h4 patterns"):
                    # Patterns
                    if d.get("daily_patterns"):
                        st.markdown("**Daily patterns:**")
                        for p in d["daily_patterns"]:
                            color = PALETTE["green"] if p.direction == "CALL" else PALETTE["red"]
                            st.markdown(f"<div style='font-family:monospace;font-size:0.78rem;color:#d4dce8;padding:2px 0;'>"
                                        f"<b style='color:{color};'>{p.direction}</b> · "
                                        f"{p.pattern} · {p.reason}</div>",
                                        unsafe_allow_html=True)

                    if d.get("h4_patterns"):
                        st.markdown("**4H patterns:**")
                        for p in d["h4_patterns"]:
                            color = PALETTE["green"] if p.direction == "CALL" else PALETTE["red"]
                            st.markdown(f"<div style='font-family:monospace;font-size:0.78rem;color:#d4dce8;padding:2px 0;'>"
                                        f"<b style='color:{color};'>{p.direction}</b> · "
                                        f"{p.pattern} · {p.reason}</div>",
                                        unsafe_allow_html=True)

                    # Gate results per direction
                    if d.get("results"):
                        st.markdown("**Gate analysis:**")
                        for r in d["results"]:
                            dir_color = PALETTE["green"] if r["direction"] == "CALL" else PALETTE["red"]
                            st.markdown(f"""
                            <div style='background:#1a1f2e;border-radius:6px;padding:10px 14px;margin:6px 0;
                                        font-family:monospace;font-size:0.78rem;'>
                              <b style='color:{dir_color};'>{r["direction"]}</b> direction:<br>
                              • GEX: {"✅" if r["gex_supports"] else "❌"} {r["gex_summary"]}<br>
                              • Whales: {"✅" if r["whale_supports"] else "❌"} {r["whale_summary"]}<br>
                              <b style='color:#bc8cff;'>→ {r["blocked_at"]}</b>
                            </div>
                            """, unsafe_allow_html=True)

    elif st.session_state.last_diagnostic is None:
        st.markdown(f"""
        <div style='text-align:center;padding:60px 0;'>
          <div style='font-size:3rem;'>🔬</div>
          <h3 style='color:{PALETTE["text"]};font-family:monospace;margin-top:12px;'>Diagnostic Mode</h3>
          <p style='color:{PALETTE["text_dim"]};font-family:monospace;font-size:0.85rem;'>
            Click <b style='color:{PALETTE["brand"]};'>Run Diagnostic Scan</b> to see where setups are getting filtered
          </p>
        </div>""", unsafe_allow_html=True)

# ════════════════════════════════════════════════
#  TAB 3 — MORNING BRIEFING
# ════════════════════════════════════════════════
with tab_morning:
    render_morning_briefing(api_key=ALPACA_KEY, api_secret=ALPACA_SECRET)


# ── Subscriber Management ─────────────────────────────────────────────────────

st.markdown("---")

with st.expander("📧 Manage Email Subscribers", expanded=False):
    from subscribers import (
        load_subscribers as _load_subs,
        add_subscriber as _add_sub,
        remove_subscriber as _rm_sub,
        is_valid_email,
    )

    current_subs = _load_subs()

    st.markdown(f"**Current subscribers: {len(current_subs)}**")
    st.caption("Emails are sent at 8 AM / 1 PM / 4 PM CT when setups are found. 4-of-4 ELITE includes Strat alignment.")

    if current_subs:
        for sub in current_subs:
            cols = st.columns([5, 1])
            with cols[0]:
                if sub["name"] != sub["email"].split("@")[0]:
                    st.markdown(
                        f"<div style='font-family:monospace;color:#d4dce8;padding:4px 0;'>"
                        f"<b style='color:#bc8cff'>{sub['name']}</b> · "
                        f"<span style='color:#a8b3c8'>{sub['email']}</span></div>",
                        unsafe_allow_html=True
                    )
                else:
                    st.markdown(
                        f"<div style='font-family:monospace;color:#d4dce8;padding:4px 0;'>"
                        f"{sub['email']}</div>",
                        unsafe_allow_html=True
                    )
            with cols[1]:
                if st.button("🗑️", key=f"rm_{sub['email']}", help="Remove"):
                    ok, msg = _rm_sub(sub["email"])
                    if ok:
                        st.success(msg)
                        st.rerun()
                    else:
                        st.error(msg)
    else:
        st.info("No subscribers yet. Add one below.")

    st.markdown("##### ➕ Add Subscriber")
    add_cols = st.columns([2, 3, 1])
    with add_cols[0]:
        new_name = st.text_input("Name (optional)", key="new_sub_name",
                                   placeholder="John", label_visibility="collapsed")
    with add_cols[1]:
        new_email = st.text_input("Email", key="new_sub_email",
                                    placeholder="john@example.com",
                                    label_visibility="collapsed")
    with add_cols[2]:
        if st.button("Add", key="add_sub_btn"):
            if not new_email:
                st.error("Email required")
            elif not is_valid_email(new_email):
                st.error("Invalid email format")
            else:
                ok, msg = _add_sub(new_email, new_name)
                if ok:
                    st.success(msg)
                    st.rerun()
                else:
                    st.warning(msg)

    st.caption(
        "⚠️ Changes are written to subscribers.txt on the deployed instance. "
        "For permanent persistence, edit subscribers.txt in GitHub and push."
    )

# Footer
st.markdown("---")
st.markdown(f"""
<div style='text-align:center;font-family:JetBrains Mono,monospace;font-size:0.68rem;color:#1a0a30;'>
  SwingConfluence · Alpaca real-time data · 3-of-3 confluence · 4-of-4 ELITE w/ Strat
</div>""", unsafe_allow_html=True)
